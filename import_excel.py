"""
import_excel.py
==============
Importa consultoras, skills e heatmap de notas do Excel para o Supabase.

Pré-requisitos:
    pip install openpyxl supabase python-dotenv

Uso:
    1. Copie este script para a mesma pasta onde está o arquivo .xlsx
    2. Crie um arquivo .env com SUPABASE_URL e SUPABASE_SERVICE_KEY
    3. Execute:  python import_excel.py

Variáveis de ambiente necessárias (.env):
    SUPABASE_URL=https://xxxx.supabase.co
    SUPABASE_SERVICE_KEY=eyJ...   (use a service_role key, não a anon key)
    EXCEL_PATH=./C&D - Cadastro de Consultoras 2026-27.xlsx

Atenção:
    - A service_role key nunca deve ser exposta no frontend.
    - Este script é apenas para uso interno/CLI de importação.
    - Execute apenas em ambiente seguro (não em produção pública).
"""

import os
import sys
import logging
from datetime import datetime, timezone
from dotenv import load_dotenv
import openpyxl
from supabase import create_client, Client

# =============================================================
# CONFIGURAÇÃO
# =============================================================

load_dotenv()

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)

SUPABASE_URL = os.environ.get("SUPABASE_URL")
SUPABASE_SERVICE_KEY = os.environ.get("SUPABASE_SERVICE_KEY")
EXCEL_PATH = os.environ.get("EXCEL_PATH", "./C&D - Cadastro de Consultoras 2026-27.xlsx")

if not SUPABASE_URL or not SUPABASE_SERVICE_KEY:
    log.error("Configure SUPABASE_URL e SUPABASE_SERVICE_KEY no arquivo .env")
    sys.exit(1)

# =============================================================
# MAPEAMENTO DE NÍVEIS TEXTUAIS → NUMÉRICO
# =============================================================

NIVEL_PARA_NOTA = {
    "mestre (faço de forma recorrente)":    3,
    "mestre":                                3,
    "praticante (fiz algumas vezes)":        2,
    "praticante":                            2,
    "aprendiz (fiz apenas 1 vez)":           1,
    "aprendiz":                              1,
    "não possuo":                            0,
    "nao possuo":                            0,
    "":                                      0,
}

def texto_para_nota(texto: str | None) -> int:
    """Converte nível textual da planilha para nota numérica 0-3."""
    if texto is None:
        return 0
    normalizado = str(texto).strip().lower()
    for chave, nota in NIVEL_PARA_NOTA.items():
        if chave in normalizado:
            return nota
    log.warning(f"  Nível não reconhecido: '{texto}' — usando 0")
    return 0


# =============================================================
# LEITURA DO EXCEL
# =============================================================

def ler_base_consultoras(wb: openpyxl.Workbook) -> dict[str, dict]:
    """
    Lê a aba 'Base' e retorna um dicionário:
    { nome_curto: { email, nome, cidade, formatos, modalidade, ... } }

    O nome_curto é como a consultora aparece no heatmap (ex: 'Daniela Gomes').
    """
    ws = wb["Base"]
    headers = [cell.value for cell in ws[1]]

    # Mapeamento de colunas da planilha
    col = {
        "id":           headers.index("Id") if "Id" in headers else None,
        "email":        headers.index("Email"),
        "nome":         headers.index("Nome"),
        "cidade":       headers.index("Cidade / Estado / País de residência"),
        "formatos":     headers.index("Em quais formatos você pode atuar?"),
        "modalidade":   None,  # não existe como coluna direta na base atual
        "viagem":       headers.index("Possui disponibilidade para viagens?"),
        "anos_exp":     headers.index("Há quantos anos você atua conduzindo treinamentos ou facilitação corporativa?"),
        "publicos":     headers.index("Com quais públicos você costuma trabalhar?"),
        "setores":      headers.index("Quais setores você já atendeu?"),
        "certificacoes":headers.index("Certificações que possui"),
        "diferencial":  headers.index("Qual é o principal diferencial da sua atuação como consultora ou facilitadora?"),
    }

    consultoras = {}
    emails_vistos = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        email = row[col["email"]]
        nome = row[col["nome"]]

        # Pular linhas sem email ou nome, e linhas de cabeçalho repetido
        if not email or not nome:
            continue
        if str(email).strip().lower() in ("email", ""):
            continue
        if email in emails_vistos:
            log.debug(f"  Email duplicado ignorado: {email}")
            continue

        emails_vistos.add(email)

        def val(campo):
            idx = col.get(campo)
            if idx is None:
                return None
            v = row[idx]
            return str(v).strip() if v else None

        consultoras[email] = {
            "nome":                 val("nome"),
            "email":                email,
            "cidade":               val("cidade"),
            "formatos":             val("formatos"),
            "disponibilidade_viagem": val("viagem"),
            "anos_experiencia":     val("anos_exp"),
            "publicos":             val("publicos"),
            "setores":              val("setores"),
            "certificacoes":        val("certificacoes"),
            "diferencial":          val("diferencial"),
            "ativo":                True,
        }

    log.info(f"Base: {len(consultoras)} consultoras com email encontradas")
    return consultoras


def ler_heatmap(wb: openpyxl.Workbook, nome_aba: str) -> tuple[list[dict], list[str]]:
    """
    Lê uma aba de heatmap e retorna:
    - lista de skills: [{ nome, origem, trilha }, ...]
    - lista de nomes de consultoras (como aparecem no cabeçalho)
    - matriz de notas: { (nome_consultora, nome_skill): nota }
    """
    ws = wb[nome_aba]
    headers = [cell.value for cell in ws[1]]

    # Colunas de skills: Tema, Origem, Trilha — colunas de consultoras a partir da 4ª
    nomes_consultoras = [h for h in headers[3:] if h is not None]

    skills = []
    notas = {}  # { (nome_consultora, nome_skill): int }

    for row in ws.iter_rows(min_row=2, values_only=True):
        tema = row[0]
        if not tema:
            continue

        origem = row[1] if len(row) > 1 else None
        trilha = row[2] if len(row) > 2 else None

        skills.append({
            "nome":    str(tema).strip(),
            "origem":  str(origem).strip() if origem else None,
            "trilha":  str(trilha).strip() if trilha else None,
        })

        for i, nome_consultora in enumerate(nomes_consultoras):
            col_idx = 3 + i
            if col_idx < len(row):
                valor = row[col_idx]
                # O heatmap já tem números; a base tem texto
                if isinstance(valor, (int, float)):
                    nota = int(valor)
                else:
                    nota = texto_para_nota(valor)
                notas[(nome_consultora, str(tema).strip())] = nota

    log.info(f"Aba '{nome_aba}': {len(skills)} skills, {len(nomes_consultoras)} consultoras")
    return skills, nomes_consultoras, notas


# =============================================================
# MAPEAMENTO NOME HEATMAP → EMAIL
# Algumas consultoras aparecem com nome curto no heatmap mas
# com nome completo na base. Ajuste aqui se necessário após
# preencher os emails faltantes de: Daniella Camara, Esteban,
# Glenda e Luis Maurício.
# =============================================================

NOME_HEATMAP_PARA_EMAIL = {
    # Mapeamentos confirmados pelos dados atuais
    "Cinthya":          "cinthya.calvo.externo@ciadetalentos.com",
    "Daniela Gomes":    "daniela.gomes@ciadetalentos.com",
    "Eliana":           "eliana.mourao@ciadetalentos.com",
    "Polyana":          "polyana.freitas.externo@ciadetalentos.com",
    "Ingrid":           "ingrid.ferreira.externo@ciadetalentos.com",

    "Daniella Camara":  "daniella.camara.externo@ciadetalentos.com",
    "Esteban":          "esteban@ciadetalentos.com",
    "Glenda":           "glenda.moreira.externo@ciadetalentos.com",
    "Luis Maurício":    "luis.mauricio@ciadetalentos.com",
}


def resolver_email(nome_heatmap: str, base: dict[str, dict]) -> str | None:
    """
    Tenta resolver o email de uma consultora a partir do nome no heatmap.
    Primeiro verifica o dicionário manual; depois tenta busca na base por nome parcial.
    """
    # Tentativa 1: mapeamento manual
    email = NOME_HEATMAP_PARA_EMAIL.get(nome_heatmap)
    if email:
        return email

    # Tentativa 2: busca por primeiro nome na base
    primeiro_nome = nome_heatmap.split()[0].lower()
    for email_base, dados in base.items():
        nome_base = dados["nome"].lower()
        if primeiro_nome in nome_base:
            log.debug(f"  Resolvido por nome parcial: {nome_heatmap} → {email_base}")
            return email_base

    log.warning(f"  ⚠️ Email não encontrado para '{nome_heatmap}' — consultora será ignorada")
    return None


# =============================================================
# IMPORTAÇÃO NO SUPABASE
# =============================================================

def upsert_consultoras(sb: Client, base: dict[str, dict]) -> dict[str, str]:
    """
    Insere ou atualiza consultoras. Retorna { email: id_uuid }.
    """
    email_para_id = {}
    registros = list(base.values())

    if not registros:
        log.warning("Nenhuma consultora para importar")
        return email_para_id

    log.info(f"Importando {len(registros)} consultoras...")
    resp = sb.table("consultoras").upsert(registros, on_conflict="email").execute()

    if resp.data:
        for item in resp.data:
            email_para_id[item["email"]] = item["id"]
        log.info(f"  ✓ {len(resp.data)} consultoras inseridas/atualizadas")
    else:
        log.error(f"  Erro ao inserir consultoras: {resp}")

    return email_para_id


def upsert_skills(sb: Client, skills: list[dict]) -> dict[str, str]:
    """
    Insere ou atualiza skills. Retorna { nome_skill: id_uuid }.
    """
    nome_para_id = {}

    if not skills:
        log.warning("Nenhuma skill para importar")
        return nome_para_id

    log.info(f"Importando {len(skills)} skills...")
    resp = sb.table("skills").upsert(skills, on_conflict="nome").execute()

    if resp.data:
        for item in resp.data:
            nome_para_id[item["nome"]] = item["id"]
        log.info(f"  ✓ {len(resp.data)} skills inseridas/atualizadas")
    else:
        log.error(f"  Erro ao inserir skills: {resp}")

    return nome_para_id


def upsert_notas(
    sb: Client,
    notas: dict[tuple, int],
    nome_heatmap_para_email: dict[str, str | None],
    email_para_id: dict[str, str],
    nome_para_id: dict[str, str],
    base: dict[str, dict],
) -> None:
    """
    Insere ou atualiza a tabela consultora_skills com as notas do heatmap.
    """
    registros = []
    ignorados = 0

    for (nome_consultora, nome_skill), nota in notas.items():
        email = resolver_email(nome_consultora, base)
        if not email:
            ignorados += 1
            continue

        consultora_id = email_para_id.get(email)
        skill_id = nome_para_id.get(nome_skill)

        if not consultora_id or not skill_id:
            log.debug(f"  ID não encontrado: {nome_consultora}/{nome_skill}")
            ignorados += 1
            continue

        nivel_texto = next(
            (k for k, v in NIVEL_PARA_NOTA.items() if v == nota and k not in ("", "nao possuo")),
            str(nota)
        )

        registros.append({
            "consultora_id": consultora_id,
            "skill_id":      skill_id,
            "nota":          nota,
            "nivel_texto":   nivel_texto,
        })

    if not registros:
        log.warning("Nenhuma nota para importar (verifique os emails pendentes)")
        return

    log.info(f"Importando {len(registros)} notas ({ignorados} ignoradas por email/id faltante)...")

    # Upsert em lotes de 200 para evitar timeout
    LOTE = 200
    total = 0
    for i in range(0, len(registros), LOTE):
        lote = registros[i:i + LOTE]
        resp = sb.table("consultora_skills").upsert(
            lote,
            on_conflict="consultora_id,skill_id"
        ).execute()
        if resp.data:
            total += len(resp.data)
        else:
            log.error(f"  Erro no lote {i//LOTE + 1}: {resp}")

    log.info(f"  ✓ {total} notas inseridas/atualizadas")


# =============================================================
# MAIN
# =============================================================

def main():
    log.info("=" * 60)
    log.info("IMPORTAÇÃO EXCEL → SUPABASE | Buscador C&D")
    log.info("=" * 60)

    # Conectar ao Supabase
    sb: Client = create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY)
    log.info("✓ Conexão Supabase estabelecida")

    # Abrir Excel
    if not os.path.exists(EXCEL_PATH):
        log.error(f"Arquivo não encontrado: {EXCEL_PATH}")
        sys.exit(1)

    log.info(f"Lendo arquivo: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH)

    # 1. Ler base de consultoras (fonte dos emails)
    base = ler_base_consultoras(wb)

    # 2. Ler heatmap principal (Jovem / Ambos)
    skills_main, _, notas_main = ler_heatmap(wb, "Heatmap (Temas vs. Consultoras)")

    # 3. Ler heatmap 1ª Liderança (se tiver dados)
    try:
        skills_lider, _, notas_lider = ler_heatmap(wb, "Heatmap (Tema 1ª Liderança)")
    except Exception:
        log.warning("Aba 'Heatmap (Tema 1ª Liderança)' não encontrada ou vazia — ignorando")
        skills_lider, notas_lider = [], {}

    # Mesclar skills (deduplicar por nome)
    skills_por_nome = {s["nome"]: s for s in skills_main}
    for s in skills_lider:
        if s["nome"] not in skills_por_nome:
            skills_por_nome[s["nome"]] = s
    todas_skills = list(skills_por_nome.values())

    # Mesclar notas
    todas_notas = {**notas_main, **notas_lider}

    # 4. Importar para o Supabase
    email_para_id = upsert_consultoras(sb, base)
    nome_para_id  = upsert_skills(sb, todas_skills)
    upsert_notas(sb, todas_notas, NOME_HEATMAP_PARA_EMAIL, email_para_id, nome_para_id, base)

    # 5. Resumo
    log.info("=" * 60)
    log.info("IMPORTAÇÃO CONCLUÍDA")
    log.info(f"  Consultoras:  {len(email_para_id)}")
    log.info(f"  Skills:       {len(nome_para_id)}")
    log.info(f"  Notas (pares): {len(todas_notas)}")

    # Alertar sobre emails pendentes
    pendentes = [
        nome for nome, email in NOME_HEATMAP_PARA_EMAIL.items() if email is None
    ]
    if pendentes:
        log.warning("")
        log.warning("⚠️  ATENÇÃO: as consultoras abaixo estão NO HEATMAP mas")
        log.warning("   SEM EMAIL configurado. Suas notas NÃO foram importadas.")
        log.warning("   Preencha NOME_HEATMAP_PARA_EMAIL neste script e rode novamente:")
        for nome in pendentes:
            log.warning(f"     → {nome}")

    log.info("=" * 60)


if __name__ == "__main__":
    main()
